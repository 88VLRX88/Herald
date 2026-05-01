from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from io import StringIO
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - exercised only when optional dep is absent
    Workbook = None
    load_workbook = None


SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
CELL_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")


def list_sheets(context: Any, path: str) -> str:
    workbook_path = workbook_file(context, path)
    wb = open_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sheets = [
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
            for sheet in wb.worksheets
        ]
    finally:
        wb.close()

    return to_json(
        {
            "workbook": relpath(context, workbook_path),
            "sheets": sheets,
        }
    )


def preview_csv(
    context: Any,
    path: str,
    sheet: str | None = None,
    max_rows: int | None = None,
    max_cols: int | None = None,
    data_only: bool = False,
    include_formulas: bool = True,
) -> str:
    workbook_path = workbook_file(context, path)
    max_rows = positive_int(max_rows, int(context.config.get("default_preview_rows", 30)))
    max_cols = positive_int(max_cols, int(context.config.get("default_preview_cols", 20)))
    wb = open_workbook(workbook_path, read_only=True, data_only=bool(data_only) if not include_formulas else False)
    value_wb = open_workbook(workbook_path, read_only=True, data_only=True) if include_formulas else None
    try:
        ws = select_sheet(wb, sheet)
        value_ws = value_wb[ws.title] if value_wb is not None else None
        rows = sheet_rows(ws, max_rows=max_rows, max_cols=max_cols, value_sheet=value_ws)
        csv_text = rows_to_csv_text(rows)
        truncated = ws.max_row > max_rows or ws.max_column > max_cols
        return to_json(
            {
                "workbook": relpath(context, workbook_path),
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "returned_rows": len(rows),
                "returned_columns": max((len(row) for row in rows), default=0),
                "truncated": truncated,
                "formula_format": formula_format_text() if include_formulas else None,
                "csv": csv_text,
            }
        )
    finally:
        wb.close()
        if value_wb is not None:
            value_wb.close()


def export_csv(
    context: Any,
    path: str,
    sheet: str | None = None,
    output_dir: str | None = None,
    data_only: bool = False,
    include_formulas: bool = True,
) -> str:
    workbook_path = workbook_file(context, path)
    target_dir = export_directory(context, workbook_path, output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    wb = open_workbook(workbook_path, read_only=True, data_only=bool(data_only) if not include_formulas else False)
    value_wb = open_workbook(workbook_path, read_only=True, data_only=True) if include_formulas else None
    exports = []
    try:
        worksheets = [select_sheet(wb, sheet)] if sheet else list(wb.worksheets)
        for ws in worksheets:
            csv_path = unique_csv_path(target_dir, ws.title)
            value_ws = value_wb[ws.title] if value_wb is not None else None
            rows_written, columns_written = write_sheet_csv(ws, csv_path, value_sheet=value_ws)
            exports.append(
                {
                    "sheet": ws.title,
                    "csv_path": relpath(context, csv_path),
                    "rows": rows_written,
                    "columns": columns_written,
                    "formula_format": formula_format_text() if include_formulas else None,
                }
            )
    finally:
        wb.close()
        if value_wb is not None:
            value_wb.close()

    return to_json(
        {
            "workbook": relpath(context, workbook_path),
            "exports": exports,
        }
    )


def set_cell(
    context: Any,
    path: str,
    sheet: str,
    cell: str,
    value: Any,
    output_path: str | None = None,
) -> str:
    workbook_path = workbook_file(context, path)
    cell = str(cell).strip().upper()
    if not CELL_RE.match(cell):
        raise ValueError(f"Invalid cell address: {cell}")

    output = output_workbook_path(context, workbook_path, output_path)
    wb = open_workbook(workbook_path, read_only=False, data_only=False)
    try:
        ws = select_sheet(wb, sheet)
        ws[cell] = value
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
    finally:
        wb.close()

    return to_json(
        {
            "workbook": relpath(context, workbook_path),
            "sheet": ws.title,
            "cell": cell,
            "value": value,
            "output_path": relpath(context, output),
        }
    )


def update_from_csv(
    context: Any,
    path: str,
    csv_path: str,
    sheet: str | None = None,
    output_path: str | None = None,
    clear_sheet: bool = True,
    infer_types: bool = True,
) -> str:
    workbook_path = workbook_file(context, path)
    source_csv = csv_file(context, csv_path)
    output = output_workbook_path(context, workbook_path, output_path)
    rows = read_csv_rows(source_csv, infer_types=bool(infer_types))

    wb = open_workbook(workbook_path, read_only=False, data_only=False)
    try:
        ws = select_or_create_sheet(wb, sheet)
        if clear_sheet:
            clear_worksheet(ws)
        for row in rows:
            ws.append(row)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
    finally:
        wb.close()

    return to_json(
        {
            "workbook": relpath(context, workbook_path),
            "csv_path": relpath(context, source_csv),
            "sheet": ws.title,
            "rows_written": len(rows),
            "columns_written": max((len(row) for row in rows), default=0),
            "output_path": relpath(context, output),
        }
    )


def create_from_csv(
    context: Any,
    csv_path: str,
    output_path: str,
    sheet: str = "Sheet1",
    infer_types: bool = True,
) -> str:
    require_openpyxl()
    source_csv = csv_file(context, csv_path)
    output = context.workspace_path(output_path)
    if output.suffix.lower() not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ValueError("output_path must end with .xlsx or .xlsm")

    rows = read_csv_rows(source_csv, infer_types=bool(infer_types))
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = str(sheet or "Sheet1")
        for row in rows:
            ws.append(row)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
    finally:
        wb.close()

    return to_json(
        {
            "csv_path": relpath(context, source_csv),
            "sheet": ws.title,
            "rows_written": len(rows),
            "columns_written": max((len(row) for row in rows), default=0),
            "output_path": relpath(context, output),
        }
    )


def require_openpyxl() -> None:
    if load_workbook is None or Workbook is None:
        raise RuntimeError("xlsx_tools requires openpyxl. Install it with: pip install openpyxl")


def open_workbook(path: Path, *, read_only: bool, data_only: bool) -> Any:
    require_openpyxl()
    return load_workbook(path, read_only=read_only, data_only=data_only)


def workbook_file(context: Any, path: str) -> Path:
    target = context.workspace_path(path)
    if not target.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if target.suffix.lower() not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ValueError("Workbook path must end with .xlsx or .xlsm")
    return target


def csv_file(context: Any, path: str) -> Path:
    target = context.workspace_path(path)
    if not target.is_file():
        raise FileNotFoundError(f"CSV not found: {path}")
    if target.suffix.lower() != ".csv":
        raise ValueError("CSV path must end with .csv")
    return target


def select_sheet(workbook: Any, sheet: str | None) -> Any:
    if not sheet:
        return workbook.active
    if sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet}. Available: {', '.join(workbook.sheetnames)}")
    return workbook[sheet]


def select_or_create_sheet(workbook: Any, sheet: str | None) -> Any:
    if not sheet:
        return workbook.active
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    return workbook.create_sheet(title=sheet)


def clear_worksheet(sheet: Any) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)


def sheet_rows(sheet: Any, *, max_rows: int, max_cols: int, value_sheet: Any | None = None) -> list[list[str]]:
    rows = []
    last_row = min(max_rows, sheet.max_row or 1)
    last_col = min(max_cols, sheet.max_column or 1)
    for row_index in range(1, last_row + 1):
        values = []
        for column_index in range(1, last_col + 1):
            formula_value = sheet.cell(row=row_index, column=column_index).value
            cached_value = value_sheet.cell(row=row_index, column=column_index).value if value_sheet is not None else None
            values.append(combined_cell_text(formula_value, cached_value))
        rows.append(trim_trailing_empty(values))
    return trim_trailing_empty_rows(rows)


def write_sheet_csv(sheet: Any, path: Path, value_sheet: Any | None = None) -> tuple[int, int]:
    rows_written = 0
    columns_written = 0
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        for row_index in range(1, (sheet.max_row or 1) + 1):
            values = []
            for column_index in range(1, (sheet.max_column or 1) + 1):
                formula_value = sheet.cell(row=row_index, column=column_index).value
                cached_value = value_sheet.cell(row=row_index, column=column_index).value if value_sheet is not None else None
                values.append(combined_cell_text(formula_value, cached_value))
            values = trim_trailing_empty(values)
            columns_written = max(columns_written, len(values))
            writer.writerow(values)
            rows_written += 1
    return rows_written, columns_written


def combined_cell_text(formula_value: Any, cached_value: Any) -> str:
    if is_formula(formula_value):
        value_text = cell_to_text(cached_value)
        formula_text = str(formula_value)[1:]
        return f"{value_text} | {formula_text}"
    return cell_to_text(formula_value)


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def formula_format_text() -> str:
    return (
        "Formula cells are exported as '<calculated value> | <formula without leading equals sign>', "
        "for example '12 | A1+B1'. If the workbook has no cached calculated value, "
        "the left side is empty."
    )


def rows_to_csv_text(rows: list[list[str]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)
    return buffer.getvalue().strip("\r\n")


def read_csv_rows(path: Path, *, infer_types: bool) -> list[list[Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        rows = []
        for row in reader:
            if infer_types:
                rows.append([infer_cell(value) for value in row])
            else:
                rows.append(row)
        return rows


def infer_cell(value: str) -> Any:
    text = value.strip()
    if text == "":
        return None
    lowered = text.lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    if lowered in {"null", "none"}:
        return None
    if re.match(r"^-?[0-9]+$", text):
        try:
            return int(text)
        except ValueError:
            return value
    if re.match(r"^-?[0-9]+\.[0-9]+$", text):
        try:
            return float(text)
        except ValueError:
            return value
    return value


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def trim_trailing_empty(values: list[str]) -> list[str]:
    while values and values[-1] == "":
        values.pop()
    return values


def trim_trailing_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    while rows and not rows[-1]:
        rows.pop()
    return rows


def positive_int(value: int | None, default: int) -> int:
    try:
        parsed = int(value) if value is not None else int(default)
    except (TypeError, ValueError):
        parsed = int(default)
    return max(1, parsed)


def export_directory(context: Any, workbook_path: Path, output_dir: str | None) -> Path:
    if output_dir:
        return context.workspace_path(output_dir)
    base = context.data_path(context.config.get("default_export_dir", "exports"))
    return base / safe_filename(workbook_path.stem)


def output_workbook_path(context: Any, workbook_path: Path, output_path: str | None) -> Path:
    if output_path:
        output = context.workspace_path(output_path)
    else:
        output = workbook_path.with_name(f"{workbook_path.stem}.edited{workbook_path.suffix}")
    if output.suffix.lower() not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ValueError("output_path must end with .xlsx or .xlsm")
    return output


def unique_csv_path(directory: Path, sheet_name: str) -> Path:
    base = safe_filename(sheet_name) or "sheet"
    candidate = directory / f"{base}.csv"
    if not candidate.exists():
        return candidate
    index = 2
    while True:
        candidate = directory / f"{base}-{index}.csv"
        if not candidate.exists():
            return candidate
        index += 1


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._-")[:80]


def relpath(context: Any, path: Path) -> str:
    try:
        return path.resolve().relative_to(context.workspace.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def to_json(value: dict[str, Any]) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2, default=cell_to_text)
