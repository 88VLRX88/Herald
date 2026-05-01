import json
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from herald_agent.runtime import Runtime
from herald_agent.tools.registry import execute_tool


ROOT = Path(__file__).resolve().parents[1]


def install_xlsx_plugin(workspace: Path) -> None:
    target = workspace / "plugins" / "xlsx_tools"
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(ROOT / "plugins" / "xlsx_tools", target)


def create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["name", "qty", "total"])
    sheet.append(["alpha", 2, "=B2+10"])
    sheet.append(["beta", 5])
    workbook.create_sheet("Notes").append(["ok"])
    workbook.save(path)
    workbook.close()
    set_cached_formula_value(path, "C2", "12")


def set_cached_formula_value(path: Path, cell: str, value: str) -> None:
    items = []
    with zipfile.ZipFile(path, "r") as source:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                text = text.replace(
                    f'<c r="{cell}"><f>B2+10</f><v></v></c>',
                    f'<c r="{cell}"><f>B2+10</f><v>{value}</v></c>',
                )
                data = text.encode("utf-8")
            items.append((item, data))
    with zipfile.ZipFile(path, "w") as target:
        for item, data in items:
            target.writestr(item, data)


def runtime(workspace: Path) -> Runtime:
    return Runtime(
        config={
            "agent": {"mode": "coder"},
            "plugins": {"directories": ["plugins"]},
        },
        workspace=workspace,
    )


def test_xlsx_plugin_previews_and_exports_csv(tmp_path: Path) -> None:
    install_xlsx_plugin(tmp_path)
    workbook_path = tmp_path / "book.xlsx"
    create_workbook(workbook_path)

    sheets = json.loads(
        execute_tool(
            {"tool": "xlsx_tools_list_sheets", "args": {"path": "book.xlsx"}},
            runtime(tmp_path),
        )
    )
    preview = json.loads(
        execute_tool(
            {"tool": "xlsx_tools_preview_csv", "args": {"path": "book.xlsx", "sheet": "Data"}},
            runtime(tmp_path),
        )
    )
    exported = json.loads(
        execute_tool(
            {"tool": "xlsx_tools_export_csv", "args": {"path": "book.xlsx", "sheet": "Data"}},
            runtime(tmp_path),
        )
    )

    assert [sheet["name"] for sheet in sheets["sheets"]] == ["Data", "Notes"]
    assert preview["formula_format"].startswith("Formula cells are exported")
    assert preview["csv"] == "name,qty,total\r\nalpha,2,12 | B2+10\r\nbeta,5"
    csv_path = tmp_path / exported["exports"][0]["csv_path"]
    assert csv_path.read_text(encoding="utf-8").splitlines() == ["name,qty,total", "alpha,2,12 | B2+10", "beta,5"]


def test_xlsx_plugin_edits_workbook_and_updates_from_csv(tmp_path: Path) -> None:
    install_xlsx_plugin(tmp_path)
    workbook_path = tmp_path / "book.xlsx"
    create_workbook(workbook_path)

    edited = json.loads(
        execute_tool(
            {
                "tool": "xlsx_tools_set_cell",
                "args": {"path": "book.xlsx", "sheet": "Data", "cell": "B2", "value": 9},
            },
            runtime(tmp_path),
        )
    )
    changed = load_workbook(tmp_path / edited["output_path"])
    try:
        assert changed["Data"]["B2"].value == 9
    finally:
        changed.close()

    csv_path = tmp_path / "replacement.csv"
    csv_path.write_text("name,qty\ngamma,11\n", encoding="utf-8")
    replaced = json.loads(
        execute_tool(
            {
                "tool": "xlsx_tools_update_from_csv",
                "args": {
                    "path": "book.xlsx",
                    "csv_path": "replacement.csv",
                    "sheet": "Data",
                    "output_path": "book.replaced.xlsx",
                },
            },
            runtime(tmp_path),
        )
    )
    workbook = load_workbook(tmp_path / replaced["output_path"])
    try:
        assert workbook["Data"]["A2"].value == "gamma"
        assert workbook["Data"]["B2"].value == 11
    finally:
        workbook.close()


def test_xlsx_plugin_creates_workbook_from_csv(tmp_path: Path) -> None:
    install_xlsx_plugin(tmp_path)
    (tmp_path / "source.csv").write_text("name,qty\nalpha,2\n", encoding="utf-8")

    created = json.loads(
        execute_tool(
            {
                "tool": "xlsx_tools_create_from_csv",
                "args": {"csv_path": "source.csv", "output_path": "created.xlsx", "sheet": "Data"},
            },
            runtime(tmp_path),
        )
    )

    workbook = load_workbook(tmp_path / created["output_path"])
    try:
        assert workbook.sheetnames == ["Data"]
        assert workbook["Data"]["A2"].value == "alpha"
        assert workbook["Data"]["B2"].value == 2
    finally:
        workbook.close()
